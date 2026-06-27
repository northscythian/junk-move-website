import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# Путь для сохранения
output_path = "D:/distr/Github projects/Canada Biss/Voicebox_Price_List_Fixed.xlsx"

# ==================== ЛИСТ 1: Junk Removal ====================
jr_data = {
    "Загрузка грузовика": ["Минимальный вызов", "¼ кузова", "½ кузова", "¾ кузова", "Полный кузов"],
    "Описание": [
        "1–2 предмета (диван, холодильник, матрас)",
        "1 комната или небольшой гараж",
        "2–3 комнаты или стандартный гараж",
        "3+ комнаты, подвал, крупная уборка",
        "Полный дом, коммерческий вывоз, стройка"
    ],
    "Цена (CAD)": [99, 150, 275, 325, 400]
}
df_jr = pd.DataFrame(jr_data)

# ==================== ЛИСТ 2: Movers (почасовые) ====================
m_hourly_data = {
    "Услуга": [
        "Команда: 2 грузчика + грузовик",
        "Дополнительный грузчик",
        "Только рабочая сила (1 чел.)",
        "Упаковка вещей (1 упаковщик)"
    ],
    "Ставка (CAD/час)": [150, 35, 70, 65]
}
df_m_hourly = pd.DataFrame(m_hourly_data)

# ==================== ЛИСТ 3: Movers (фиксированные) ====================
m_flat_data = {
    "Тип недвижимости": [
        "1-комнатная / студия",
        "2-комнатная",
        "3-комнатная",
        "Небольшой дом (~100 м²)",
        "Большой дом (100–200 м²)"
    ],
    "Средняя цена (CAD)": [450, 1000, 1500, 2000, 3000]
}
df_m_flat = pd.DataFrame(m_flat_data)

# ==================== ЛИСТ 4: Combo Packages ====================
combo_data = {
    "Пакет": ["Лёгкий переезд", "Стандартный переезд", "Полный сервис"],
    "Junk Removal": ["¼ кузова", "½ кузова", "Полный кузов"],
    "Movers": ["1-комнатная", "2-комнатная", "Дом 100–200 м²"],
    "Цена (CAD)": [625, 1275, 2900],
    "Экономия клиента": [75, 125, 200]
}
df_combo = pd.DataFrame(combo_data)

# ==================== ЛИСТ 5: Profit Calculator ====================
profit_data = {
    "Показатель": ["Выручка (за день)", "Себестоимость (50%)", "Маркетинг (12%)", "Накладные (12%)", "ЧИСТАЯ ПРИБЫЛЬ"],
    "Значение (CAD)": [1000, 500, 120, 120, 260],
    "% от выручки": ["100%", "50%", "12%", "12%", "26%"]
}
df_profit = pd.DataFrame(profit_data)

# ==================== СОХРАНЯЕМ ВСЕ ЛИСТЫ В ОДИН ФАЙЛ ====================
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df_jr.to_excel(writer, sheet_name='Junk Removal', index=False)
    df_m_hourly.to_excel(writer, sheet_name='Movers (Hourly)', index=False)
    df_m_flat.to_excel(writer, sheet_name='Movers (Flat Rate)', index=False)
    df_combo.to_excel(writer, sheet_name='Combo Packages', index=False)
    df_profit.to_excel(writer, sheet_name='Profit Calculator', index=False)

# ==================== ДОБАВЛЯЕМ ФОРМАТИРОВАНИЕ (опционально) ====================
wb = load_workbook(output_path)

# Стили
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# Применяем стили ко всем листам
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Заголовки
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    
    # Данные
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 1:
                cell.alignment = left_align
            elif col == ws.max_column:
                cell.alignment = right_align
            
            # Формат денег для последнего столбца (если это числа)
            if col == ws.max_column and isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.00'
    
    # Ширина колонок
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[chr(64 + col)].width = min(max_len + 5, 50)

wb.save(output_path)
print(f"✅ Файл успешно создан: {output_path}")